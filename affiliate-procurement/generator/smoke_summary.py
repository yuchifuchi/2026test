# -*- coding: utf-8 -*-
"""02_比較評価シート に記入例データを流し込み、03_稟議用サマリ の
順位・選定候補順位・稟議文面が正しく出るかを実データで確認する。"""
import sys
import formulas
from openpyxl import load_workbook

from build import build_example, R_ITEM0, R_ITEM_LAST, R_TCO_SCORE, R_COST0, VCOLS

SRC = sys.argv[1]
TMP = "smoke_filled.xlsx"

ex = build_example()
wb = load_workbook(SRC)
ws = wb["02_比較評価シート"]

for j, name in enumerate(ex["vendors"]):
    ws[f"{VCOLS[j]}8"] = name if name else None
for j, costs in enumerate(ex["cost"]):
    for i, v in enumerate(costs):
        ws[f"{VCOLS[j]}{R_COST0+i}"] = v
for j, scores in enumerate(ex["scores"]):
    for i, v in enumerate(scores):
        r = R_ITEM0 + i
        if r == R_TCO_SCORE:
            continue
        ws[f"{VCOLS[j]}{r}"] = v
ws["B3"] = ex["header"][0]
wb.calculation.fullCalcOnLoad = True
wb.save(TMP)

xl = formulas.ExcelModel().loads(TMP).finish()
sol = xl.calculate()
cells = {}
for k, v in sol.items():
    if "!" not in k:
        continue
    ref = k.split("]", 1)[-1]
    try:
        val = v.value[0, 0]
    except Exception:
        try:
            val = v.value.ravel()[0]
        except Exception:
            val = v
    cells[ref] = val


def g(ref):
    return cells.get(f"03_稟議用サマリ'!{ref}")


errs = [k for k, v in cells.items()
        if any(e in str(v) for e in ("#REF!", "#VALUE!", "#NAME?", "#DIV/0!", "#N/A", "#NUM!"))]
print("エラー値:", len(errs), errs[:10])

print("\n--- 03_稟議用サマリ 総合評価(記入例データ投入時)---")
print(f"{'ベンダー':<26}{'Must':<10}{'総合点':>8}{'点数順位':>8}{'選定候補':>8}{'TCO':>14}{'TCO順':>6}")
for r in range(6, 10):
    name = g(f"C{r}")
    print(f"{str(name):<26}{str(g(f'D{r}')):<10}"
          f"{(round(g(f'E{r}'),1) if isinstance(g(f'E{r}'),(int,float)) else g(f'E{r}')):>8}"
          f"{str(g(f'F{r}')):>8}{str(g(f'G{r}')):>8}"
          f"{(f'{int(g(chr(72)+str(r))):,}' if isinstance(g(f'H{r}'),(int,float)) else g(f'H{r}')):>14}"
          f"{str(g(f'I{r}')):>6}")

print("\n--- 稟議記載用たたき台 ---")
for r in range(44, 49):
    v = g(f"C{r}")
    if v:
        print("  ", v)

print("\n--- 大分類別スコア ---")
for r in range(13, 19):
    print("  ", g(f"B{r}"), "w=", g(f"C{r}"),
          [round(g(f"{c}{r}"), 2) if isinstance(g(f"{c}{r}"), (int, float)) else g(f"{c}{r}")
           for c in "DEFG"])
