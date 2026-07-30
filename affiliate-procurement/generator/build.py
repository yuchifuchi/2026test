# -*- coding: utf-8 -*-
"""業務システム調達 ベンダー比較評価シート を生成する。"""
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import RadarChart, Reference

from items import CATEGORIES, PRESETS, ITEMS, QUESTIONS

FONT = "Meiryo"

NAVY = "1F3864"
LIGHT_BLUE = "D9E2F3"
INPUT_YELLOW = "FFF2CC"
AUTO_GRAY = "EDEDED"
BAND = "F7F9FC"
ACCENT = "C00000"

thin = Side(style="thin", color="B4C6E7")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

SHEET_EVAL = "02_比較評価シート"
SHEET_WEIGHT = "01_重み付け設定"
SHEET_SUM = "03_稟議用サマリ"
SHEET_EX = "04_記入例"

# 評価グリッドの位置(両シート共通)
R_VENDOR = 8                      # ベンダー名入力行
R_COST0 = 12                      # 費用入力の先頭行
R_COST_LAST = 17                  # 費用入力の最終行
R_TCO = 18                        # 5年TCO行
R_ITEM0 = 25                      # 評価項目の先頭行
R_ITEM_LAST = R_ITEM0 + len(ITEMS) - 1
R_TCO_SCORE = R_ITEM0 + 6         # TCO自動採点の行(7番目の項目)
VCOLS = ["F", "G", "H", "I"]

COST_ROWS = [
    ("初期費用(ライセンス・構築)", "一時"),
    ("カスタマイズ・追加開発費", "一時"),
    ("データ移行費用", "一時"),
    ("自社側の追加コスト(端末・回線・外部委託等)", "一時"),
    ("年間ランニング費用(保守・利用料)", "年額"),
    ("年間その他費用(法改正対応・VUP等の想定)", "年額"),
]


def cell(ws, ref, value=None, *, bold=False, size=10, color="000000",
         fill=None, align=None, wrap=False, fmt=None, border=False, italic=False):
    c = ws[ref]
    if value is not None:
        c.value = value
    c.font = Font(name=FONT, size=size, bold=bold, color=color, italic=italic)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = BORDER
    return c


def title_bar(ws, ref_range, text):
    ws.merge_cells(ref_range)
    first = ref_range.split(":")[0]
    cell(ws, first, text, bold=True, size=13, color="FFFFFF", fill=NAVY, align="left")
    ws.row_dimensions[int("".join(ch for ch in first if ch.isdigit()))].height = 26


def section(ws, ref_range, text):
    ws.merge_cells(ref_range)
    first = ref_range.split(":")[0]
    cell(ws, first, text, bold=True, size=11, color=NAVY, fill=LIGHT_BLUE, align="left")


def build_eval_sheet(ws, example=False, ex=None):
    """評価シート本体。example=True なら記入例データを流し込む。"""
    widths = {"A": 14, "B": 5, "C": 34, "D": 52, "E": 6,
              "F": 13, "G": 13, "H": 13, "I": 13, "J": 40}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    head = "【記入例】業務システム調達 ベンダー比較評価シート" if example \
        else "業務システム調達 ベンダー比較評価シート"
    title_bar(ws, "A1:J1", head)

    if example:
        cell(ws, "A2",
             "※ 架空案件の記入例です。実際の評価には 02_比較評価シート を使ってください。",
             size=9, color=ACCENT, italic=True)

    for i, label in enumerate(["案件名", "評価者", "評価日"]):
        r = 3 + i
        cell(ws, f"A{r}", label, bold=True, size=10, align="right")
        ws.merge_cells(f"B{r}:E{r}")
        v = ex["header"][i] if example else None
        cell(ws, f"B{r}", v, fill=None if example else INPUT_YELLOW, border=True, align="left")

    # ---- STEP1 ベンダー名 ----
    section(ws, "A7:J7", "■ STEP1  比較するベンダー名を入力する(最大4社)")
    ws.merge_cells(f"A{R_VENDOR}:E{R_VENDOR}")
    cell(ws, f"A{R_VENDOR}", "ベンダー名 →", bold=True, align="right")
    for i, col in enumerate(VCOLS):
        v = (ex["vendors"][i] if example else f"{chr(65+i)}社")
        cell(ws, f"{col}{R_VENDOR}", v, bold=True, align="center",
             fill=None if example else INPUT_YELLOW, border=True)
    cell(ws, f"J{R_VENDOR}",
         "※ ここに入れた社名が全シートに反映されます", size=9, color="808080")

    # ---- STEP2 費用 ----
    section(ws, "A10:J10", "■ STEP2  費用を入力する(5年総額 TCO を自動算出)")
    cell(ws, "C11", "費用項目(税抜・円)", bold=True, fill=LIGHT_BLUE, border=True, align="center")
    cell(ws, "D11", "区分", bold=True, fill=LIGHT_BLUE, border=True, align="center")
    for col in VCOLS:
        cell(ws, f"{col}11", f'=IF({col}{R_VENDOR}="","",{col}{R_VENDOR})', bold=True, fill=LIGHT_BLUE,
             border=True, align="center")
    cell(ws, "J11", "メモ", bold=True, fill=LIGHT_BLUE, border=True, align="center")

    for i, (name, kind) in enumerate(COST_ROWS):
        r = R_COST0 + i
        cell(ws, f"C{r}", name, border=True, wrap=True, align="left")
        cell(ws, f"D{r}", kind, border=True, align="center", size=9)
        for j, col in enumerate(VCOLS):
            v = ex["cost"][j][i] if (example and j < len(ex["cost"])) else None
            cell(ws, f"{col}{r}", v, fill=None if example else INPUT_YELLOW,
                 border=True, fmt="#,##0", align="right")
        cell(ws, f"J{r}", (ex["cost_memo"][i] if example else None),
             border=True, wrap=True, size=9, align="left")

    cell(ws, f"C{R_TCO}", "■ 5年総額(TCO)", bold=True, fill=AUTO_GRAY, border=True)
    cell(ws, f"D{R_TCO}", "自動", bold=True, fill=AUTO_GRAY, border=True, align="center", size=9)
    for col in VCOLS:
        f = (f'=IF(COUNT({col}{R_COST0}:{col}{R_COST_LAST})=0,"",'
             f'SUM({col}{R_COST0}:{col}{R_COST0+3})+SUM({col}{R_COST0+4}:{col}{R_COST_LAST})*5)')
        cell(ws, f"{col}{R_TCO}", f, bold=True, fill=AUTO_GRAY, border=True,
             fmt="#,##0", align="right")
    cell(ws, f"J{R_TCO}", "一時費用の合計 + 年額の合計×5年", border=True, size=9, color="808080")

    cell(ws, "C19",
         "※ 初期費用の安さで判断しないこと。TCOで並べ替えると順位が入れ替わるのが普通です。",
         size=9, color=ACCENT)

    # ---- STEP3 評価 ----
    section(ws, "A21:J21", "■ STEP3  各項目を 1〜5 で評価する")
    cell(ws, "A22",
         "5 = 期待を大きく上回る  /  4 = 十分に満たす  /  3 = 標準的・条件付きで満たす  "
         "/  2 = 不足あり、対応策が必要  /  1 = 満たさない",
         size=9, color="404040")
    cell(ws, "A23",
         "Must 欄が「○」の項目で 2 以下が1つでもあると、総合点に関わらず「要再検討」と判定されます。",
         size=9, color=ACCENT)

    headers = ["大分類", "No", "評価項目", "確認の観点(ここを見る)", "Must"]
    for i, h in enumerate(headers):
        cell(ws, f"{get_column_letter(i+1)}24", h, bold=True, color="FFFFFF",
             fill=NAVY, border=True, align="center", wrap=True)
    for col in VCOLS:
        cell(ws, f"{col}24", f'=IF({col}{R_VENDOR}="","",{col}{R_VENDOR})', bold=True, color="FFFFFF",
             fill=NAVY, border=True, align="center")
    cell(ws, "J24", "メモ・確認結果", bold=True, color="FFFFFF", fill=NAVY,
         border=True, align="center")
    ws.row_dimensions[24].height = 30

    for i, (cat, name, view, must) in enumerate(ITEMS):
        r = R_ITEM0 + i
        band = BAND if i % 2 else None
        cell(ws, f"A{r}", cat, border=True, size=9, align="center", wrap=True, fill=band)
        cell(ws, f"B{r}", i + 1, border=True, size=9, align="center", fill=band)
        cell(ws, f"C{r}", name, border=True, wrap=True, align="left", fill=band)
        cell(ws, f"D{r}", view, border=True, wrap=True, align="left", size=9,
             color="404040", fill=band)
        cell(ws, f"E{r}", must if must else "-", border=True, align="center",
             bold=bool(must), color=ACCENT if must else "808080", fill=band)
        ws.row_dimensions[r].height = 34

        if r == R_TCO_SCORE:
            for col in VCOLS:
                f = (f'=IF(COUNT($F${R_TCO}:$I${R_TCO})=0,"",IF({col}${R_TCO}="","",'
                     f'ROUND(MAX(1,MIN(5,5*MIN($F${R_TCO}:$I${R_TCO})/{col}${R_TCO})),1)))')
                cell(ws, f"{col}{r}", f, border=True, align="center", fill=AUTO_GRAY,
                     fmt="0.0", bold=True)
        else:
            for j, col in enumerate(VCOLS):
                v = None
                if example and j < len(ex["scores"]):
                    v = ex["scores"][j][i]
                cell(ws, f"{col}{r}", v, border=True, align="center",
                     fill=band if example else INPUT_YELLOW, bold=True)
        memo = ex["memo"][i] if example else None
        cell(ws, f"J{r}", memo, border=True, wrap=True, size=9, align="left", fill=band)

    # 1〜5 の入力規則(自動算出行は除く)
    dv = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True,
                        showErrorMessage=True, errorTitle="入力エラー",
                        error="1〜5 の整数で入力してください。")
    ws.add_data_validation(dv)
    dv.add(f"F{R_ITEM0}:I{R_TCO_SCORE-1}")
    dv.add(f"F{R_TCO_SCORE+1}:I{R_ITEM_LAST}")

    # ---- 大分類別スコア(集計の土台) ----
    r0 = R_ITEM_LAST + 2
    rc0 = r0 + 5
    cell(ws, f"C{rc0-1}", "大分類別スコア(5点満点)  ※自動計算", bold=True, size=9, color="808080")
    for i, (cid, cname, w, _d) in enumerate(CATEGORIES):
        r = rc0 + i
        cell(ws, f"C{r}", cname, border=True, size=9, fill=AUTO_GRAY)
        cell(ws, f"D{r}", f"='{SHEET_WEIGHT}'!C{5+i}", border=True, size=9,
             align="center", fill=AUTO_GRAY, fmt="0")
        for col in VCOLS:
            f = (f'=IFERROR(AVERAGEIFS({col}${R_ITEM0}:{col}${R_ITEM_LAST},'
                 f'$A${R_ITEM0}:$A${R_ITEM_LAST},$C{r}),0)')
            cell(ws, f"{col}{r}", f, border=True, size=9, align="center",
                 fmt="0.00", fill=AUTO_GRAY)

    # ---- 集計 ----
    section(ws, f"A{r0}:J{r0}", "■ 集計結果(順位・チャート・稟議文面は 03_稟議用サマリ へ)")
    labels = ["Must判定", "加重総合点(100点満点)", "5年総額(TCO)"]
    for i, lb in enumerate(labels):
        r = r0 + 1 + i
        cell(ws, f"C{r}", lb, bold=True, fill=AUTO_GRAY, border=True)
        for col in VCOLS:
            if i == 0:
                f = (f'=IF(COUNT({col}{R_ITEM0}:{col}{R_ITEM_LAST})=0,"未評価",'
                     f'IF(SUMPRODUCT(($E${R_ITEM0}:$E${R_ITEM_LAST}="○")*'
                     f'({col}${R_ITEM0}:{col}${R_ITEM_LAST}<3)*'
                     f'({col}${R_ITEM0}:{col}${R_ITEM_LAST}>0))>0,"要再検討","合格"))')
                cell(ws, f"{col}{r}", f, border=True, align="center", bold=True, fill=AUTO_GRAY)
            elif i == 1:
                f = (f'=IF(COUNT({col}{R_ITEM0}:{col}{R_ITEM_LAST})=0,"",'
                     f'SUMPRODUCT($D${rc0}:$D${rc0+5},{col}${rc0}:{col}${rc0+5})/100/5*100)')
                cell(ws, f"{col}{r}", f, border=True, align="center", bold=True,
                     fmt="0.0", fill=AUTO_GRAY)
            else:
                cell(ws, f"{col}{r}", f"={col}{R_TCO}", border=True, align="right",
                     fmt="#,##0", fill=AUTO_GRAY)

    ws.freeze_panes = "C25"
    ws.sheet_view.showGridLines = False
    return rc0


def main(out):
    wb = Workbook()

    # ============ 00_使い方 ============
    ws = wb.active
    ws.title = "00_使い方"
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 4, "B": 30, "C": 86}.items():
        ws.column_dimensions[col].width = w
    title_bar(ws, "A1:C1", "業務システム調達 ベンダー比較評価シート  ─ 使い方")

    lines = [
        ("", ""),
        ("■ このシートの目的", ""),
        ("", "比較表を作ることではなく、ベンダーに会う前に「評価軸」と「優先順位」を確定させることです。"),
        ("", "評価軸を後から作ると、提案書の上手いベンダーが勝ちます。順序を逆にしないでください。"),
        ("", ""),
        ("■ 進め方", ""),
        ("STEP 0", "01_重み付け設定 で、6つの大分類に重みを割り振る。プリセットを選んで微調整するのが早い。"),
        ("", "ここはベンダーに会う前に決めきる。会った後に重みを変えると、それは評価ではなく後付けの理由づけです。"),
        ("STEP 1", "02_比較評価シート にベンダー名を入力する(最大4社)。全シートに自動反映されます。"),
        ("STEP 2", "費用を入力する。5年総額(TCO)が自動算出され、コストの点数も自動でつきます。"),
        ("STEP 3", "29項目を 1〜5 で評価する。分からない項目は空欄のまま、05_ベンダー質問リスト で確認する。"),
        ("STEP 4", "03_稟議用サマリ で総合点・順位・レーダーチャート・稟議文面のたたき台を確認する。"),
        ("", ""),
        ("■ 色の意味", ""),
        ("黄色のセル", "入力欄。ここだけを埋めてください。"),
        ("灰色のセル", "自動計算。触ると壊れます。"),
        ("", ""),
        ("■ Must(必須要件)の扱い", ""),
        ("", "Must欄が「○」の項目で 2以下 が1つでもあると、総合点に関わらず「要再検討」と判定されます。"),
        ("", "点数だけで運用すると、必須要件を満たさない会社が総合点で勝ってしまう事故が起きます。"),
        ("", "この分離が、このシートで最も重要な仕組みです。判定を無視して進める場合は、理由を稟議に明記してください。"),
        ("", ""),
        ("■ 使うときの注意", ""),
        ("1", "評点は必ず根拠とセットで。メモ欄が空の点数は、3か月後の自分が説明できません。"),
        ("2", "利用部門の担当者にも同じシートで評価してもらい、点数の差が大きい項目を議論の起点にする。"),
        ("3", "初期費用ではなくTCOで並べる。順位が入れ替わるのが普通です。"),
        ("4", "「自社側に必要な工数」は見積書に出てきません。必ずベンダーに人日で出させてください。"),
        ("5", "評価項目・重みは自社に合わせて足し引きして構いません。ただしベンダーに会う前に確定させること。"),
        ("", ""),
        ("■ 免責", ""),
        ("", "本シートは調達判断を補助するツールであり、特定の製品・ベンダーの選定結果を保証するものではありません。"),
        ("", "契約・法務・会計上の判断は、必ず自社の該当部門および専門家にご確認ください。"),
    ]
    r = 2
    for label, text in lines:
        if label.startswith("■"):
            ws.merge_cells(f"B{r}:C{r}")
            cell(ws, f"B{r}", label, bold=True, size=11, color=NAVY, fill=LIGHT_BLUE)
        else:
            cell(ws, f"B{r}", label, bold=True, size=10, align="right", color="404040")
            cell(ws, f"C{r}", text, size=10, wrap=True, align="left")
            if text:
                ws.row_dimensions[r].height = 20
        r += 1

    # ============ 01_重み付け設定 ============
    ws = wb.create_sheet(SHEET_WEIGHT)
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 4, "B": 18, "C": 12, "D": 62, "E": 15, "F": 15, "G": 15, "H": 15}.items():
        ws.column_dimensions[col].width = w
    title_bar(ws, "A1:H1", "STEP 0  評価軸の重み付け設定")
    cell(ws, "B2", "合計が100になるように配分してください。右のプリセットを転記して微調整するのが最短です。",
         size=10, color="404040")
    cell(ws, "B3", "重みは必ずベンダーに会う前に確定させること。後から変えると評価ではなく後付けの理由づけになります。",
         size=9, color=ACCENT)

    hdr = ["", "大分類", "重み(%)", "この軸で見るもの"] + [p[0] for p in PRESETS]
    for i, h in enumerate(hdr):
        cell(ws, f"{get_column_letter(i+1)}4", h, bold=True, color="FFFFFF",
             fill=NAVY, border=True, align="center", wrap=True)
    ws.row_dimensions[4].height = 34

    for i, (cid, name, w, desc) in enumerate(CATEGORIES):
        r = 5 + i
        cell(ws, f"A{r}", cid, border=True, size=9, align="center", color="808080")
        cell(ws, f"B{r}", name, border=True, bold=True)
        cell(ws, f"C{r}", w, border=True, align="center", bold=True,
             fill=INPUT_YELLOW, fmt="0")
        cell(ws, f"D{r}", desc, border=True, wrap=True, size=9, color="404040", align="left")
        for j, (_pn, vals) in enumerate(PRESETS):
            cell(ws, f"{get_column_letter(5+j)}{r}", vals[i], border=True,
                 align="center", size=9, fill=BAND)
        ws.row_dimensions[r].height = 30

    r = 5 + len(CATEGORIES)
    cell(ws, f"B{r}", "合計", bold=True, fill=AUTO_GRAY, border=True, align="right")
    cell(ws, f"C{r}", f"=SUM(C5:C{r-1})", bold=True, fill=AUTO_GRAY, border=True,
         align="center", fmt="0")
    cell(ws, f"D{r}", f'=IF(C{r}=100,"OK  合計100です","※ 合計が100になっていません('
                      f'"&C{r}&")。調整してください")',
         bold=True, border=True, fill=AUTO_GRAY, align="left")
    for j in range(len(PRESETS)):
        col = get_column_letter(5 + j)
        cell(ws, f"{col}{r}", f"=SUM({col}5:{col}{r-1})", fill=AUTO_GRAY,
             border=True, align="center", size=9)

    cell(ws, "B" + str(r + 2), "重み付けの決め方", bold=True, size=11, color=NAVY)
    tips = [
        "・「全部大事」は重み付けではありません。必ず差をつけてください。差がつかない配分は判断を先送りしているだけです。",
        "・迷ったら「この案件が失敗するとしたら、原因は何か」を考え、その軸を重くします。",
        "・現場の反対で頓挫しそうなら機能適合、情シスの手が足りないなら導入負荷とサポート、が実務的な目安です。",
        "・重みは利用部門・上長と合意しておくと、稟議での説明がそのまま通ります。",
    ]
    for i, t in enumerate(tips):
        ws.merge_cells(f"B{r+3+i}:H{r+3+i}")
        cell(ws, f"B{r+3+i}", t, size=9, color="404040", align="left")

    # ============ 02_比較評価シート ============
    ws_eval = wb.create_sheet(SHEET_EVAL)
    rc0 = build_eval_sheet(ws_eval, example=False)

    # ============ 03_稟議用サマリ ============
    ws = wb.create_sheet(SHEET_SUM)
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 3, "B": 5, "C": 20, "D": 12, "E": 18, "F": 10,
                   "G": 14, "H": 16, "I": 10, "J": 3, "K": 3, "L": 12}.items():
        ws.column_dimensions[col].width = w
    title_bar(ws, "A1:I1", "STEP 4  稟議用サマリ")
    cell(ws, "B2", "案件名:", bold=True, align="right")
    cell(ws, "C2", f"='{SHEET_EVAL}'!B3", bold=True, size=11)

    section(ws, "B4:I4", "■ 総合評価")
    sh = ["#", "ベンダー", "Must判定", "加重総合点(100点)", "点数順位",
          "選定候補順位", "5年総額(円)", "TCO順位"]
    for i, h in enumerate(sh):
        cell(ws, f"{get_column_letter(2+i)}5", h, bold=True, color="FFFFFF",
             fill=NAVY, border=True, align="center", wrap=True)
    ws.row_dimensions[5].height = 32

    for j in range(4):
        r = 6 + j
        col = VCOLS[j]
        cell(ws, f"B{r}", j + 1, border=True, align="center", size=9)
        cell(ws, f"C{r}", f'=IF(\'{SHEET_EVAL}\'!{col}{R_VENDOR}="","",\'{SHEET_EVAL}\'!{col}{R_VENDOR})', border=True, bold=True)
        cell(ws, f"D{r}", f"='{SHEET_EVAL}'!{col}{R_ITEM_LAST+3}", border=True,
             align="center", bold=True)
        cell(ws, f"E{r}",
             f'=IF(SUM(${get_column_letter(4+j)}$13:${get_column_letter(4+j)}$18)=0,"",'
             f'SUMPRODUCT($C$13:$C$18,{get_column_letter(4+j)}$13:{get_column_letter(4+j)}$18)/100/5*100)',
             border=True, align="center", bold=True, fmt="0.0")
        cell(ws, f"F{r}", f'=IF($E{r}="","-",RANK($E{r},$E$6:$E$9,0))',
             border=True, align="center")
        cell(ws, f"G{r}", f'=IF($L{r}="","-",RANK($L{r},$L$6:$L$9,0))',
             border=True, align="center", bold=True)
        cell(ws, f"H{r}", f"='{SHEET_EVAL}'!{col}{R_TCO}", border=True,
             align="right", fmt="#,##0")
        cell(ws, f"I{r}", f'=IF(ISNUMBER($H{r}),RANK($H{r},$H$6:$H$9,1),"-")',
             border=True, align="center")
        cell(ws, f"L{r}", f'=IF(AND($D{r}="合格",ISNUMBER($E{r})),$E{r},"")',
             size=8, color="BFBFBF", fmt="0.0", align="center")

    cell(ws, "L5", "(補助列)", size=8, color="BFBFBF", align="center")
    cell(ws, "B10",
         "※「選定候補順位」は Must判定が「合格」のベンダーのみを対象にした順位です。点数順位と食い違う場合、"
         "見るべきはこちらです。",
         size=9, color=ACCENT)

    section(ws, "B11:I11", "■ 大分類別スコア(5点満点)")
    for j in range(4):
        cell(ws, f"{get_column_letter(4+j)}12", f'=IF(\'{SHEET_EVAL}\'!{VCOLS[j]}{R_VENDOR}="","",\'{SHEET_EVAL}\'!{VCOLS[j]}{R_VENDOR})',
             bold=True, color="FFFFFF", fill=NAVY, border=True, align="center")
    cell(ws, "B12", "大分類", bold=True, color="FFFFFF", fill=NAVY, border=True, align="center")
    cell(ws, "C12", "重み(%)", bold=True, color="FFFFFF", fill=NAVY, border=True, align="center")

    for i, (_cid, cname, _w, _d) in enumerate(CATEGORIES):
        r = 13 + i
        cell(ws, f"B{r}", f"='{SHEET_WEIGHT}'!B{5+i}", border=True, size=10)
        cell(ws, f"C{r}", f"='{SHEET_WEIGHT}'!C{5+i}", border=True, align="center", fmt="0")
        for j in range(4):
            cell(ws, f"{get_column_letter(4+j)}{r}",
                 f"='{SHEET_EVAL}'!{VCOLS[j]}{rc0+i}", border=True,
                 align="center", fmt="0.00")

    chart = RadarChart()
    chart.type = "marker"
    chart.style = 26
    chart.title = "大分類別スコア比較"
    data = Reference(ws, min_col=4, max_col=7, min_row=12, max_row=18)
    cats = Reference(ws, min_col=2, min_row=13, max_row=18)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height, chart.width = 10, 13
    ws.add_chart(chart, "B21")

    r = 42
    section(ws, f"B{r}:I{r}", "■ 稟議記載用のたたき台(空欄を埋めて転記してください)")
    top = 'INDEX($C$6:$C$9,MATCH(1,$G$6:$G$9,0))'
    tops = 'INDEX($E$6:$E$9,MATCH(1,$G$6:$G$9,0))'
    topt = 'INDEX($H$6:$H$9,MATCH(1,$G$6:$G$9,0))'
    body = [
        ("【選定結果】", None),
        (None, f'="評価の結果、"&IFERROR({top},"(未評価)")&" を選定候補とする。"'),
        ("【選定理由】", None),
        (None, '="1. 必須要件(Must項目)をすべて満たしていることを確認した。"'),
        (None, f'="2. 6つの評価軸に重み付けを行った加重評価で "&IFERROR(TEXT({tops},"0.0"),"-")'
               f'&" 点(100点満点)を獲得し、Must合格社の中で最上位である。"'),
        (None, f'="3. 5年総額(TCO)は "&IFERROR(TEXT({topt},"#,##0"),"-")&" 円(税抜)。"'),
        (None, '="4. (定性的な理由をここに追記)"'),
        ("【他社を選定しなかった理由】", None),
        (None, "(ここに記入。落選各社について1〜2行ずつ、決め手となった項目名を挙げて書くと通りやすい)"),
        ("【想定リスクと対応】", None),
        (None, "(ここに記入。評価で 2 以下をつけた項目が、そのままリスク欄の材料になります)"),
        ("【比較検討の経緯】", None),
        (None, "(ここに記入。ベンダー選定の母数、RFP提示日、提案受領日、デモ実施日などを時系列で)"),
    ]
    for i, (label, formula) in enumerate(body):
        rr = r + 1 + i
        if label:
            cell(ws, f"B{rr}", label, bold=True, color=NAVY, size=10)
        else:
            ws.merge_cells(f"C{rr}:I{rr}")
            cell(ws, f"C{rr}", formula, size=10, align="left", wrap=True,
                 fill=None if str(formula).startswith("=") else INPUT_YELLOW,
                 border=not str(formula).startswith("="))
            ws.row_dimensions[rr].height = 18

    ws.freeze_panes = "B6"

    # ============ 04_記入例 ============
    ex = build_example()
    ws_ex = wb.create_sheet(SHEET_EX)
    rc0_ex = build_eval_sheet(ws_ex, example=True, ex=ex)
    r_read = rc0_ex + 8
    cell(ws_ex, f"C{r_read}", "▼ この記入例の読みどころ", bold=True, size=11, color=NAVY)
    reads = [
        "・B社は加重総合点で1位だが、必須要件を標準機能で満たせず(項目1で2点)Must判定が「要再検討」。"
        "点数だけで決めていたら、要件を満たさない製品を選んでいたことになります。",
        "・C社は機能適合が最も高いが、拡張性・将来性が 2.6 と低い。スクラッチ開発の典型で、"
        "作った直後は最適でも5年後に動けなくなる形です。",
        "・一時費用はC社が最も高い(1,450万円 / A社1,350万円)のに、5年TCOではC社2,250万円 < A社2,350万円 と逆転します。"
        "初期費用だけで並べると順位を読み違える、という典型例です。",
        "・結果として、Must合格社の中で最上位のA社が実質の選定候補になります。",
        "・メモ欄の書き方に注目してください。点数の根拠が1行あるだけで、3か月後の稟議説明が成立します。",
    ]
    for i, t in enumerate(reads):
        rr = r_read + 1 + i
        ws_ex.merge_cells(f"C{rr}:J{rr}")
        cell(ws_ex, f"C{rr}", t, size=9, color="404040", align="left", wrap=True)
        ws_ex.row_dimensions[rr].height = 26

    # ============ 05_ベンダー質問リスト ============
    ws = wb.create_sheet("05_ベンダー質問リスト")
    ws.sheet_view.showGridLines = False
    for col, w in {"A": 4, "B": 16, "C": 62, "D": 56, "E": 8, "F": 8, "G": 8, "H": 8}.items():
        ws.column_dimensions[col].width = w
    title_bar(ws, "A1:H1", "ベンダーへの質問リスト(RFP・提案依頼時にそのまま使えます)")
    cell(ws, "B2",
         "評価項目を埋めるために必要な質問を並べています。提案依頼時に添付するか、デモ・ヒアリングの場で使ってください。",
         size=10, color="404040")
    cell(ws, "B3",
         "重要なのは「答えられるか」だけでなく「文書で出せるか」です。口頭回答は、確認済みに数えないでください。",
         size=9, color=ACCENT)

    hdr = ["No", "分類", "質問内容", "なぜ聞くか(回答をどう使うか)"]
    for i, h in enumerate(hdr):
        cell(ws, f"{get_column_letter(1+i)}5", h, bold=True, color="FFFFFF",
             fill=NAVY, border=True, align="center")
    for j in range(4):
        cell(ws, f"{get_column_letter(5+j)}5", f'=IF(\'{SHEET_EVAL}\'!{VCOLS[j]}{R_VENDOR}="","",\'{SHEET_EVAL}\'!{VCOLS[j]}{R_VENDOR})',
             bold=True, color="FFFFFF", fill=NAVY, border=True, align="center")
    ws.row_dimensions[5].height = 30

    dv2 = DataValidation(type="list", formula1='"済,未,N/A"', allow_blank=True)
    ws.add_data_validation(dv2)
    for i, (cat, q, why) in enumerate(QUESTIONS):
        r = 6 + i
        band = BAND if i % 2 else None
        cell(ws, f"A{r}", i + 1, border=True, align="center", size=9, fill=band)
        cell(ws, f"B{r}", cat, border=True, size=9, align="center", wrap=True, fill=band)
        cell(ws, f"C{r}", q, border=True, wrap=True, align="left", fill=band)
        cell(ws, f"D{r}", why, border=True, wrap=True, align="left", size=9,
             color="404040", fill=band)
        for j in range(4):
            cell(ws, f"{get_column_letter(5+j)}{r}", None, border=True,
                 align="center", fill=INPUT_YELLOW)
        ws.row_dimensions[r].height = 32
    dv2.add(f"E6:H{5+len(QUESTIONS)}")
    ws.freeze_panes = "C6"

    # openpyxl はキャッシュ値を書かないため、Excel 側で開いた瞬間に全再計算させる
    wb.calculation.fullCalcOnLoad = True
    wb.save(out)
    print("saved:", out)
    print("items:", len(ITEMS), "rows", R_ITEM0, "-", R_ITEM_LAST,
          "| tco_score_row", R_TCO_SCORE, "| cat_block", rc0)


def build_example():
    """架空案件『販売管理システム刷新(社員120名・製造業)』の記入例。"""
    vendors = ["A社(大手パッケージ)", "B社(中堅SaaS)", "C社(地元SIer/スクラッチ)", ""]
    cost = [
        [8000000, 3500000, 1200000, 800000, 1800000, 200000],
        [1500000,  800000,  900000, 600000, 2400000,      0],
        [12000000,      0, 1500000, 1000000, 1200000, 400000],
    ]
    cost_memo = [
        "A社は初年度ライセンス一括。B社は初期設定費のみ",
        "A社は帳票9本の改修。C社はスクラッチのため初期費用に内包",
        "3社とも旧システムからのマスタ移行を含む",
        "端末更新・回線増強・移行時の派遣要員。ベンダー見積には含まれない",
        "A社は保守料、B社は月額利用料×12、C社は保守料",
        "法改正対応・VUP時の想定。A社は保守に一部含む、C社は都度有償のため厚めに計上",
    ]
    #        A社 B社 C社
    S = [
        (5, 2, 5), (4, 4, 2), (5, 3, 5), (3, 3, 5), (3, 5, 3), (4, 2, 5),   # 機能適合
        (None, None, None),                                                  # TCO(自動)
        (4, 5, 3), (4, 4, 3), (3, 4, 4),                                     # コスト
        (3, 4, 4), (2, 5, 2), (3, 4, 3), (4, 4, 3), (3, 4, 3),               # 導入・移行
        (4, 3, 5), (5, 4, 3), (4, 4, 3), (3, 3, 5), (4, 3, 4),               # サポート
        (4, 5, 2), (4, 5, 3), (5, 4, 3), (3, 5, 2), (4, 4, 3),               # 拡張性
        (5, 3, 4), (3, 4, 5), (5, 3, 3), (4, 4, 5),                          # 信頼性
    ]
    memo = [
        "A社/C社は標準で充足。B社は3機能が未対応で、うち1つは代替運用も現場が拒否 → 2点",
        "C社はスクラッチのため標準機能という概念がなく、全量が個別開発。VUPリスクを考慮して2点",
        "B社は任意要件のうち4件が未対応。ただし足切り対象外",
        "A社/B社はパッケージに業務を合わせる必要あり。C社は現行踏襲。現場合意はA社で調整中",
        "利用部門3名で実機評価。B社が最も操作回数が少ない。A社は受注入力に2画面必要",
        "B社は帳票追加が都度有償(1本8万円〜)。既存帳票の再現に難あり",
        None,
        "B社は月額の内訳が明快。C社は「開発一式」表記が多く、内訳提示を再依頼した",
        "全社とも文書で回答あり。C社は法改正対応が都度見積で、金額レンジの提示を拒否 → 3点",
        "B社は年契約で解約金なし。A社は5年縛り、C社は保守3年縛り",
        "移行対象は得意先・商品・在庫の3マスタ。B社はCSV取込ツールあり。A社は移行ツールが有償",
        "A社は要員確保の都合で着手が4か月後。繁忙期(3月)にテストが重なる → 2点",
        "自社工数の見積:A社 45人日 / B社 28人日 / C社 40人日。全社から人日で回答を取得済み",
        "A社は研修2回まで無償。C社は1回のみで追加は有償",
        "B社のみ部門単位の段階切替が可能。A社/C社は一斉切替で切り戻し期限は当日中",
        "C社は地元で往訪対応可。B社はチャット中心で電話は上位プランのみ",
        "A社はSLA文書あり(一次回答2時間)。C社はSLA文書がなく、契約書に追記を要請中 → 3点",
        "C社は「操作質問は有償」の線引きが曖昧。A社/B社は範囲一覧を文書で受領済み → C社3点",
        "C社は担当が当社業務を熟知。ただし1名依存で、交代時の体制に不安あり",
        "A社/C社は四半期報告あり。B社は自動レポートのみ",
        "B社はAPI公開・仕様書あり。C社はAPIなし、CSV手動連携のみ → 2点",
        "B社は管理画面から全件エクスポート可。C社はDB直参照で、契約書に持ち出し条項なし → 追記要請中",
        "A社はインボイス対応を施行3か月前にリリース済み。C社は都度有償かつ実績の提示なし → 3点",
        "C社はスクラッチのため改修が都度有償。B社は自動VUPで自社負担ほぼなし",
        "5年後180名想定で再見積を取得。B社は段階課金で+38%、A社は+15%",
        "A社は同業種同規模の事例を5件提示。B社は業種違いが多い",
        "C社の提案書はヒアリング内容が具体的に反映されていた。A社は汎用資料の流用が目立つ → 3点",
        "A社は上場企業。B社/C社は非上場で、EOL方針は口頭回答のみ",
        "C社は往訪、B社は電話ヒアリングに応諾。A社は書面事例のみ",
    ]
    scores = [[], [], []]
    for tup in S:
        for j in range(3):
            scores[j].append(tup[j])
    return {
        "header": ["販売管理システム刷新(製造業・社員120名)", "情報システム課 ◯◯", "2026-06-15"],
        "vendors": vendors,
        "cost": cost,
        "cost_memo": cost_memo,
        "scores": scores,
        "memo": memo,
    }


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "out.xlsx")
