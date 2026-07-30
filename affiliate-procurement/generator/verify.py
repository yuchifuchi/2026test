# -*- coding: utf-8 -*-
"""LibreOffice に依存しない静的検証 + 記入例の期待値を Python で独立計算する。"""
import re
import sys
from collections import Counter
from openpyxl import load_workbook

from items import CATEGORIES, ITEMS, PRESETS
from build import build_example, R_ITEM0, R_ITEM_LAST, R_TCO_SCORE, R_TCO, R_COST0, R_COST_LAST

# 単語境界で判定する(AVERAGEIFS を IFS と誤検出しないため)
BANNED = ["XLOOKUP", "XMATCH", "SORT", "FILTER", "UNIQUE", "SEQUENCE",
          "TEXTJOIN", "CONCAT", "IFS", "SWITCH", "MAXIFS", "MINIFS"]
ALLOWED_FUNCS = {"IF", "IFERROR", "SUM", "SUMPRODUCT", "AVERAGEIFS", "COUNT",
                 "MIN", "MAX", "ROUND", "RANK", "INDEX", "MATCH", "TEXT",
                 "AND", "OR", "ISNUMBER"}

fails = []


def check_static(path):
    wb = load_workbook(path)
    funcs = Counter()
    n = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if isinstance(c.value, str) and c.value.startswith("="):
                    n += 1
                    f = c.value
                    used = re.findall(r"(?<![A-Z0-9_.])([A-Z][A-Z0-9_.]*)\s*\(", f.upper())
                    for m in used:
                        funcs[m] += 1
                        if m in BANNED:
                            fails.append(f"BANNED {m} at {ws.title}!{c.coordinate}: {f}")
    print(f"formulas: {n}")
    print("functions used:", dict(funcs))
    unknown = set(funcs) - ALLOWED_FUNCS
    if unknown:
        fails.append(f"UNEXPECTED FUNCTIONS: {unknown}")

    # マージ済みセルへの書き込み事故がないか
    for ws in wb.worksheets:
        merged = set()
        for rng in ws.merged_cells.ranges:
            for row in ws[rng.coord]:
                for c in row:
                    merged.add(c.coordinate)
            merged.discard(rng.coord.split(":")[0])
        for coord in merged:
            c = ws[coord]
            if getattr(c, "value", None) not in (None, ""):
                fails.append(f"WRITE INTO MERGED CELL {ws.title}!{coord} = {c.value!r}")

    names = [ws.title for ws in wb.worksheets]
    print("sheets:", names)
    return wb


def expected():
    """記入例の期待値を独立に計算する(シートの数式とは別ロジック)。"""
    ex = build_example()
    weights = {c[1]: c[2] for c in CATEGORIES}
    tco = []
    for v in ex["cost"]:
        tco.append(sum(v[0:4]) + sum(v[4:6]) * 5)

    n_v = len(ex["scores"])
    scores = [list(s) for s in ex["scores"]]
    idx_tco = 6
    assert R_TCO_SCORE - R_ITEM0 == idx_tco, "TCO自動採点行の位置がズレている"
    for j in range(n_v):
        s = round(max(1, min(5, 5 * min(tco) / tco[j])), 1)
        scores[j][idx_tco] = s

    print("\n--- 記入例の期待値 ---")
    print("vendors:", ex["vendors"][:n_v])
    print("5年TCO :", [f"{t:,}" for t in tco])
    print("TCO点  :", [scores[j][idx_tco] for j in range(n_v)])

    cat_avg = []
    for j in range(n_v):
        d = {}
        for cname in weights:
            vals = [scores[j][i] for i, it in enumerate(ITEMS) if it[0] == cname]
            d[cname] = sum(vals) / len(vals)
        cat_avg.append(d)
        print(f"  {ex['vendors'][j]}: " +
              " / ".join(f"{k}={v:.2f}" for k, v in d.items()))

    totals = []
    for j in range(n_v):
        t = sum(cat_avg[j][k] * weights[k] for k in weights) / 100 / 5 * 100
        totals.append(t)
    print("加重総合点:", [f"{t:.1f}" for t in totals])

    must_idx = [i for i, it in enumerate(ITEMS) if it[3] == "○"]
    print("Must項目 No.:", [i + 1 for i in must_idx])
    verdict = []
    for j in range(n_v):
        bad = [(i + 1, ITEMS[i][1], scores[j][i]) for i in must_idx
               if scores[j][i] is not None and scores[j][i] < 3]
        verdict.append("要再検討" if bad else "合格")
        if bad:
            print(f"  {ex['vendors'][j]} → 要再検討: {bad}")
    print("Must判定:", verdict)

    order = sorted(range(n_v), key=lambda j: -totals[j])
    print("点数順位:", [ex["vendors"][j] for j in order])
    cand = [j for j in order if verdict[j] == "合格"]
    print("選定候補順位:", [ex["vendors"][j] for j in cand])
    print("→ 稟議たたき台が拾うベンダー:", ex["vendors"][cand[0]] if cand else "なし")

    # 重み合計チェック
    assert sum(weights.values()) == 100, "既定重みの合計が100でない"
    for name, vals in PRESETS:
        assert sum(vals) == 100, f"プリセット {name} の合計が100でない"
    print("重み合計チェック: 既定/プリセット4種ともOK")

    # 費用行数の整合
    assert R_COST_LAST - R_COST0 + 1 == 6, "費用入力行数が想定と違う"
    assert R_ITEM_LAST - R_ITEM0 + 1 == len(ITEMS), "評価項目行数が想定と違う"
    print("行レイアウト整合: OK")


if __name__ == "__main__":
    check_static(sys.argv[1])
    expected()
    print()
    if fails:
        print("=== FAIL ===")
        for f in fails:
            print(" -", f)
        sys.exit(1)
    print("=== 静的検証 OK ===")
