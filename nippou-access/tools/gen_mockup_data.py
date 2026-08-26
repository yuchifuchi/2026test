# -*- coding: utf-8 -*-
"""モックアップ用のデータを JSON で吐く。マスタと実績はすべて現行 Excel の実データ。"""
import csv, io, json, os, sys
import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(HERE, "data")

def load(n):
    with io.open(os.path.join(DATA, n), encoding="utf-8") as f:
        return list(csv.DictReader(f))

def main(src):
    cols = {int(r["集計列ID"]): r["集計列名"] for r in load("M_集計列.csv")}
    blocks = {int(r["ブロックID"]): r["ブロック名"] for r in load("M_ブロック.csv")}
    prods = [{"id": int(r["製品ID"]), "b": int(r["ブロックID"]), "n": r["製品名"]}
             for r in load("M_製品.csv") if r["有効"] == "1"]
    prods.insert(0, {"id": 0, "b": 0, "n": "（製品指定なし）"})
    kubun = [{"id": int(r["区分ID"]), "b": int(r["ブロックID"]), "n": r["区分名"],
              "c": int(r["集計列ID"]), "old": r["旧転記名"]}
             for r in load("M_区分.csv") if r["有効"] == "1"]
    ops = [{"id": int(r["担当者ID"]), "code": r["担当者コード"], "sei": r["姓"],
            "name": r["氏名"], "kbn": r["職員区分"], "note": r["備考"] if "備考" in r else "",
            "bad": r["名前設定ミス"] == "1", "on": r["有効"] == "1"}
           for r in load("M_担当者.csv")]
    tasks = [{"id": int(r["業務項目ID"]), "n": r["帳票表示名"]} for r in load("M_業務項目.csv")]

    by_sei = {o["sei"]: o["id"] for o in ops}
    by_name = {o["name"]: o["id"] for o in ops}
    kb_by_old = {}
    for k in kubun:
        kb_by_old.setdefault((k["old"], k["c"]), k["id"])
        kb_by_old.setdefault(("", k["c"]) if not k["old"] else (k["old"], None), k["id"])
    pr_by_name = {}
    for p in prods:
        pr_by_name.setdefault((p["n"], p["b"]), p["id"])
        pr_by_name.setdefault((p["n"], None), p["id"])
    kb_block = {k["id"]: k["b"] for k in kubun}

    rows = []
    ws = openpyxl.load_workbook(os.path.join(src, "集計表（ＶＢＡ版）.xlsm"), data_only=True)["データ"]
    for r in range(9, ws.max_row + 1):
        dt = ws.cell(r, 1).value
        if not hasattr(dt, "year"):
            continue
        col, cnt = 0, 0
        for i in range(3, 9):
            v = ws.cell(r, i).value
            if isinstance(v, (int, float)) and v:
                col, cnt = i, int(v); break
        if not col:
            continue
        op = str(ws.cell(r, 11).value or "").strip()
        kb_name = str(ws.cell(r, 10).value or "").strip()
        pn = str(ws.cell(r, 2).value or "").strip()
        op_id = by_sei.get(op) or by_name.get(op) or 0
        kb_id = kb_by_old.get((kb_name, col)) or kb_by_old.get((kb_name, None)) or 0
        if not kb_id and not kb_name:
            kb_id = next((k["id"] for k in kubun if k["c"] == col and not k["old"]), 0)
        pr_id = 0
        if pn and kb_id:
            pr_id = pr_by_name.get((pn, kb_block.get(kb_id))) or pr_by_name.get((pn, None)) or 0
        if op_id and kb_id:
            rows.append({"d": dt.strftime("%Y-%m-%d"), "o": op_id, "k": kb_id,
                         "p": pr_id, "c": cnt})

    # 同じキーはまとめる (Access 版の UNIQUE 制約と同じ挙動)
    merged = {}
    for r in rows:
        key = (r["d"], r["o"], r["k"], r["p"])
        merged[key] = merged.get(key, 0) + r["c"]
    rows = [{"d": k[0], "o": k[1], "k": k[2], "p": k[3], "c": v} for k, v in merged.items()]

    out = {"cols": cols, "blocks": blocks, "prods": prods, "kubun": kubun,
           "ops": ops, "tasks": tasks, "juden": rows,
           "nippou": {"2026-08-25": {"kaisen": 5,
                                     "tokki": "昭和100年記念貨幣の抽選結果について、"
                                              "発表日の問合せが集中した。",
                                     "daitai": "支払方法の変更依頼 1 件を職員が対応。",
                                     "youbou": "抽選結果の発表日をハガキにも明記して"
                                               "いただけると、問合せが減ると思われます。"}}}
    print(json.dumps(out, ensure_ascii=False, separators=(",", ":")))

if __name__ == "__main__":
    main(sys.argv[1])
