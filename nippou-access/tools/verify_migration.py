# -*- coding: utf-8 -*-
"""現行の集計表 Excel を、Access 版と同じ規則で SQLite に取り込んで検算する。

Windows / Access を使わずに、
  ・マスタの引き当てが何行成功して何行落ちるか
  ・Access 版の集計値が Excel の SUMIF と一致するか
を先に確かめるためのもの。modImportExcel.bas と同じ引き当て規則を実装している。
"""
import argparse
import csv
import io
import os
import sqlite3
import sys
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA = os.path.join(HERE, "data")


def load_csv(name):
    with io.open(os.path.join(DATA, name), encoding="utf-8") as f:
        return list(csv.DictReader(f))


def build_db():
    db = sqlite3.connect(":memory:")
    db.executescript("""
        CREATE TABLE M_集計列 (集計列ID INT PRIMARY KEY, 集計列名 TEXT);
        CREATE TABLE M_製品   (製品ID INT PRIMARY KEY, ブロックID INT, 製品名 TEXT, 有効 INT);
        CREATE TABLE M_区分   (区分ID INT PRIMARY KEY, ブロックID INT, 区分名 TEXT,
                               集計列ID INT, 内訳区分 TEXT, 有効 INT, 旧転記名 TEXT);
        CREATE TABLE M_担当者 (担当者ID INT PRIMARY KEY, 担当者コード TEXT, 姓 TEXT,
                               氏名 TEXT, 職員区分 TEXT);
        CREATE TABLE T_受電   (対象日 TEXT, 担当者ID INT, 区分ID INT, 製品ID INT, 件数 INT,
                               UNIQUE(対象日, 担当者ID, 区分ID, 製品ID));
    """)
    for r in load_csv("M_集計列.csv"):
        db.execute("INSERT INTO M_集計列 VALUES (?,?)", (int(r["集計列ID"]), r["集計列名"]))
    db.execute("INSERT INTO M_製品 VALUES (0,0,'（製品指定なし）',1)")
    for r in load_csv("M_製品.csv"):
        db.execute("INSERT INTO M_製品 VALUES (?,?,?,?)",
                   (int(r["製品ID"]), int(r["ブロックID"]), r["製品名"], int(r["有効"])))
    for r in load_csv("M_区分.csv"):
        legacy = r["旧転記名"]
        uchi = {"返品": "返品", "返金": "返金"}.get(legacy, "")
        db.execute("INSERT INTO M_区分 VALUES (?,?,?,?,?,?,?)",
                   (int(r["区分ID"]), int(r["ブロックID"]), r["区分名"],
                    int(r["集計列ID"]), uchi, int(r["有効"]), legacy))
    for r in load_csv("M_担当者.csv"):
        db.execute("INSERT INTO M_担当者 VALUES (?,?,?,?,?)",
                   (int(r["担当者ID"]), r["担当者コード"], r["姓"], r["氏名"], r["職員区分"]))
    db.commit()
    return db


def one(db, sql, args=()):
    row = db.execute(sql, args).fetchone()
    return row[0] if row else None


# --- modImportExcel.bas と同じ引き当て規則 -----------------------------------
def find_operator(db, nm):
    nm = (nm or "").strip()
    if not nm:
        return 0
    for sql in ("SELECT 担当者ID FROM M_担当者 WHERE 姓=?",
                "SELECT 担当者ID FROM M_担当者 WHERE 氏名=?"):
        v = one(db, sql, (nm,))
        if v:
            return v
    if len(nm) > 1:
        v = one(db, "SELECT 担当者ID FROM M_担当者 WHERE 姓 LIKE ?", ("%" + nm[1:] + "%",))
        if v:
            return v
    return 0


def find_kubun(db, nm, col_id):
    nm = (nm or "").strip()
    if nm:
        for sql, args in (
            ("SELECT 区分ID FROM M_区分 WHERE 旧転記名=? AND 集計列ID=?", (nm, col_id)),
            ("SELECT 区分ID FROM M_区分 WHERE 区分名=? AND 集計列ID=?", (nm, col_id)),
            ("SELECT 区分ID FROM M_区分 WHERE 旧転記名=?", (nm,)),
        ):
            v = one(db, sql, args)
            if v:
                return v
        return 0
    return one(db, "SELECT 区分ID FROM M_区分 WHERE 集計列ID=? AND (旧転記名 IS NULL OR 旧転記名='')",
               (col_id,)) or 0


def find_product(db, nm, kb_id):
    nm = (nm or "").strip()
    if not nm:
        return 0
    if kb_id:
        blk = one(db, "SELECT ブロックID FROM M_区分 WHERE 区分ID=?", (kb_id,))
        if blk is not None:
            v = one(db, "SELECT 製品ID FROM M_製品 WHERE 製品名=? AND ブロックID=?", (nm, blk))
            if v:
                return v
    return one(db, "SELECT 製品ID FROM M_製品 WHERE 製品名=?", (nm,)) or 0


def import_sheet(db, path):
    ws = openpyxl.load_workbook(path, data_only=True)["データ"]
    ok, ng, skip = 0, 0, 0
    failures = defaultdict(int)
    excel_rows = []

    for r in range(9, ws.max_row + 1):
        dt = ws.cell(r, 1).value
        prod = (ws.cell(r, 2).value or "")
        kubun = (ws.cell(r, 10).value or "")
        op = (ws.cell(r, 11).value or "")

        col_id, cnt = 0, 0
        for i in range(3, 9):                       # C..H = 集計列 3..8
            v = ws.cell(r, i).value
            if isinstance(v, (int, float)) and v:
                col_id, cnt = i, int(v)
                break
        if not hasattr(dt, "year") or col_id == 0:
            skip += 1
            continue

        day = dt.strftime("%Y-%m-%d")
        excel_rows.append((day, col_id, cnt))       # Excel 側の生データ (検算用)

        op_id = find_operator(db, str(op))
        kb_id = find_kubun(db, str(kubun), col_id)
        pr_id = find_product(db, str(prod), kb_id)

        if op_id == 0 or kb_id == 0:
            ng += 1
            why = []
            if op_id == 0:
                why.append("担当者『%s』" % op)
            if kb_id == 0:
                why.append("区分『%s』(集計列%d)" % (kubun, col_id))
            failures[" / ".join(why)] += 1
            continue

        cur = db.execute(
            "SELECT 件数 FROM T_受電 WHERE 対象日=? AND 担当者ID=? AND 区分ID=? AND 製品ID=?",
            (day, op_id, kb_id, pr_id)).fetchone()
        if cur:
            db.execute("UPDATE T_受電 SET 件数=件数+? WHERE 対象日=? AND 担当者ID=? "
                       "AND 区分ID=? AND 製品ID=?", (cnt, day, op_id, kb_id, pr_id))
        else:
            db.execute("INSERT INTO T_受電 VALUES (?,?,?,?,?)",
                       (day, op_id, kb_id, pr_id, cnt))
        ok += 1

    db.commit()
    return ok, ng, skip, failures, excel_rows


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True)
    args = ap.parse_args()

    db = build_db()
    path = os.path.join(args.src, "集計表（ＶＢＡ版）.xlsm")
    if not os.path.exists(path):
        sys.exit("見つかりません: " + path)

    ok, ng, skip, failures, excel_rows = import_sheet(db, path)
    print("=== 取込結果 (%s) ===" % os.path.basename(path))
    print("  成功 %d 行 / 引き当て失敗 %d 行 / 対象外 %d 行" % (ok, ng, skip))
    if failures:
        print("  失敗の内訳:")
        for why, n in sorted(failures.items(), key=lambda x: -x[1]):
            print("    %3d 行  %s" % (n, why))

    print()
    print("=== 検算: Excel の生データ vs Access 版のクエリ ===")
    # Excel 側: 取り込む前の行を日付 × 集計列 で単純合計したもの
    expect = defaultdict(int)
    for day, col_id, cnt in excel_rows:
        expect[(day, col_id)] += cnt              # Excel の C..H 列 = 集計列 3..8

    # Access 版: Q_週次集計 と同じ SQL
    got = defaultdict(int)
    for day, col, cnt in db.execute(
        "SELECT J.対象日, K.集計列ID, SUM(J.件数) FROM T_受電 J "
        "JOIN M_区分 K ON J.区分ID=K.区分ID GROUP BY J.対象日, K.集計列ID"):
        got[(day, col)] = cnt

    names = dict(db.execute("SELECT 集計列ID, 集計列名 FROM M_集計列"))
    days = sorted({d for d, _ in list(expect) + list(got)})
    hdr = "  %-12s" % "日付" + "".join("%10s" % names[c] for c in range(3, 9)) + "%8s" % "計"
    print(hdr)
    print("  " + "-" * (len(hdr) - 2))
    diff_total = 0
    for day in days:
        e = [expect[(day, c)] for c in range(3, 9)]
        g = [got[(day, c)] for c in range(3, 9)]
        mark = "" if e == g else "   ← 差異"
        if e != g:
            diff_total += 1
        print("  %-12s" % day + "".join("%10d" % v for v in g) + "%8d" % sum(g) + mark)
        if e != g:
            print("  %-12s" % "  (Excel)" + "".join("%10d" % v for v in e) + "%8d" % sum(e))

    print()
    # 担当者ごとの取込結果。現行の設定ミスがどう見えるかを可視化する。
    print("=== 担当者別の取込結果 ===")
    for nm, code, n, c in db.execute(
        "SELECT OP.氏名, OP.担当者コード, COUNT(*), SUM(J.件数) FROM T_受電 J "
        "JOIN M_担当者 OP ON J.担当者ID=OP.担当者ID "
        "GROUP BY OP.氏名, OP.担当者コード ORDER BY SUM(J.件数) DESC"):
        note = ""
        if code == "100":
            note = ("   ← 現行の設定ミスで 002堀/004谷口/006国森/009西岡/015鈴木 の"
                    "5 名分が混ざっている。手で振り分けが必要")
        print("  %-4s %-12s %4d 行 %5d 件%s" % (code, nm, n, c, note))
    missing = [r[0] for r in db.execute(
        "SELECT 担当者コード || ' ' || 氏名 FROM M_担当者 "
        "WHERE 担当者ID NOT IN (SELECT DISTINCT 担当者ID FROM T_受電) ORDER BY 担当者ID")]
    if missing:
        print("  取込データが 1 行も無い担当者: " + " / ".join(missing))

    # 姓だけでは一意に決まらない担当者を警告する。
    # 現行の集計表には姓しか入っていないので、ここが曖昧だと別人に付く。
    dup = list(db.execute(
        "SELECT 姓, COUNT(*), GROUP_CONCAT(担当者コード || ':' || 氏名, ' / ') "
        "FROM M_担当者 GROUP BY 姓 HAVING COUNT(*) > 1"))
    ambiguous = list(db.execute(
        "SELECT A.姓, A.担当者コード, B.担当者コード FROM M_担当者 A, M_担当者 B "
        "WHERE A.担当者ID <> B.担当者ID AND B.姓 LIKE A.姓 || '%'"))
    if dup or ambiguous:
        print()
        print("  [注意] 姓だけでは一意に決まらない組み合わせがあります。")
        for sei, n, who in dup:
            print("    姓『%s』が %d 名: %s" % (sei, n, who))
        for sei, a, b in ambiguous:
            print("    『%s』(%s) は %s の姓の前方一致にもなります。"
                  "旧データの『%s』がどちらの分かは Excel からは判別できません。"
                  % (sei, a, b, sei))

    print()
    if diff_total == 0:
        print("  全日付で一致。差異が出るとすれば引き当て失敗行のみ。")
    else:
        print("  %d 日で差異あり (引き当て失敗行が原因)。" % diff_total)

    # 参考: Excel の SUMIF が実際に返していた値との比較
    print()
    print("=== 参考: 現行 Excel の SUMIF が返していた値 ===")
    wb = openpyxl.load_workbook(path, data_only=True)["データ"]
    for r in range(2, 7):
        d = wb.cell(r, 2).value
        if not hasattr(d, "year"):
            continue
        vals = [wb.cell(r, c).value or 0 for c in range(3, 9)]
        day = d.strftime("%Y-%m-%d")
        true_vals = [got[(day, c)] for c in range(3, 9)]
        mark = "" if list(vals) == true_vals else "   ← Excel の SUMIF が過少/過大"
        print("  %-12s SUMIF=%s  正=%s%s" % (day, vals, true_vals, mark))
    return 0


if __name__ == "__main__":
    sys.exit(main())
