# -*- coding: utf-8 -*-
"""現行 Excel マクロ一式から Access 用マスタデータを抽出する。

入力: 記入用フォーム (*.xlsm) 一式 と 日報集計印刷用フォーム
出力: data/*.csv (UTF-8) と src/modSetupMaster.bas (Access へ投入する VBA)

現行の情報源:
  - 記入用フォーム Sheet1 の見出しは 日報集計管理用フォーム!集計用フォーム への外部リンク。
    その「キャッシュ値」が xl/externalLinks/externalLinkN.xml に残っているので、そこから見出しを復元する。
  - 記入用フォーム Sheet2 の 2行目 = 集計表データシートの列番号、3行目 = 集計表へ書き込む備考名。
    Sheet2 の各セルの数式 (=+Sheet1!C6 など) が Sheet1 セル ⇔ 区分 の対応表になっている。
"""
import argparse
import csv
import os
import re
import sys
import zipfile
import xml.etree.ElementTree as ET

import openpyxl
from openpyxl.utils import get_column_letter

NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

# 集計表「データ」シートの列番号 -> 集計列名 (C..H 列)
SHUKEI_COLUMNS = [
    (3, "申込方法"),
    (4, "抽選結果"),
    (5, "納付書発送"),
    (6, "商品発送"),
    (7, "その他"),
    (8, "商品交換"),
]

# 入力画面のブロック定義。
#   (ブロックID, ブロック名, 製品別か, 見出し行, データ行 or None(=製品別),
#    見出しを組み立てる行のリスト, 対象列)
BLOCKS = [
    (1, "区分",             True,  [3, 4, 5], None,     range(6, 29),  "CDEFGHIJKLM"),
    (2, "その他",           True,  [31],      None,     range(32, 53), "CDEFGHIJKLM"),
    (3, "顧客情報",         False, [53],      54,       None,          "CDEFGHIJKLM"),
    (4, "イベント関係",     False, [55],      56,       None,          "CDEFGHIJKLM"),
    (5, "その他のその他①", False, [57],      58,       None,          "CDEFGHIJKLM"),
    (6, "その他のその他②", False, [59],      60,       None,          "CDEFGHIJKLM"),
]

# 電話応対以外の業務。(番号, 集計用フォームのラベルセル, 件数セル, 帳票上の表示名)
TASKS = [
    ("①", "A67", "C67", "①　受注入力"),
    ("②", "A68", "C68", "②　受注チェック"),
    ("③", "A69", "C69", "③　戻り郵便処理／払込書チェック"),
    ("④", "A70", "C70", "④　ＤＭ処理／戻り郵便"),
    ("⑤", "E66", "G66", "⑤　架電"),
    ("⑥", "E67", "G67", "⑥　その他（　　　　　　　　）"),
    ("⑦", "E68", "G68", "⑦　エクセル入力"),
    ("⑧", "E69", "G69", "⑧　エクセルチェック"),
    ("⑨", "E70", "G70", "⑨　アンケート入力"),
    ("⑩", "I66", "L66", "⑩　ﾊｶﾞｷﾃﾞｰﾀ入力"),
    ("⑪", "I67", "L67", "⑪　ハガキデータチェック"),
    ("⑫", "I68", "L68", "⑫　新規ｺｰﾄﾞ取り"),
    ("⑬", "I69", "L69", "⑬　顧客整理"),
]

# 特殊な問合せ (ブロック7)。件数セルは Sheet1!G62/G63、内容は Sheet1!A64。
SPECIAL = [
    ("別添のとおり", "G62"),
    ("下記のとおり（別添不要）", "G63"),
]


def fold(s):
    """全角英数を半角に畳んで比較する (顧客Ｇ と 顧客G を同一視するため)。"""
    return "".join(
        chr(ord(ch) - 0xFEE0) if 0xFF01 <= ord(ch) <= 0xFF5E else ch
        for ch in (s or "")
    ).upper()


def norm(v):
    """セル値を表示用文字列に正規化する。全角空白と改行は詰める。"""
    if v is None:
        return ""
    s = str(v).replace("　", " ").replace("\n", "").strip()
    return re.sub(r"\s+", " ", s)


def read_external_cache(path):
    """記入用フォームの外部リンクキャッシュから 集計用フォーム の見出しを取り出す。

    外部リンクは [1] と [2] の 2 本あり、ファイルによって新旧パスの向きが逆転している。
    どちらでも良いので「集計用フォーム」シートのキャッシュが最も充実している方を採用する。
    """
    z = zipfile.ZipFile(path)
    best = {}
    for entry in z.namelist():
        if not re.fullmatch(r"xl/externalLinks/externalLink\d+\.xml", entry):
            continue
        root = ET.fromstring(z.read(entry))
        book = root.find("m:externalBook", NS)
        if book is None:
            continue
        names = [sn.get("val") for sn in book.find("m:sheetNames", NS)]
        data_set = book.find("m:sheetDataSet", NS)
        if data_set is None:
            continue
        for sheet in data_set:
            idx = int(sheet.get("sheetId"))
            if idx >= len(names) or names[idx] != "集計用フォーム":
                continue
            cells = {}
            for row in sheet.findall("m:row", NS):
                for cell in row.findall("m:cell", NS):
                    val = cell.find("m:v", NS)
                    if val is not None and val.text:
                        cells[cell.get("r")] = val.text
            if len(cells) > len(best):
                best = cells
    if not best:
        raise SystemExit(f"{path}: 集計用フォームの外部リンクキャッシュが見つかりません")
    return best


def merged_fill(path, cache, header_rows):
    """見出し行の結合セルを展開する。

    集計用フォームでは「払込用紙」が F3:I3、「再発行」が F4:G4 のように横結合されており、
    キャッシュには左上セルの値しか入らない。記入用フォーム Sheet1 は同じレイアウトなので、
    その結合範囲を使って結合先の列にも同じ見出しを配る。
    """
    ws = openpyxl.load_workbook(path)["Sheet1"]
    filled = dict(cache)
    for rng in ws.merged_cells.ranges:
        origin = f"{get_column_letter(rng.min_col)}{rng.min_row}"
        val = cache.get(origin)
        if not val:
            continue
        # K3:L4 のように行方向にも広がる結合があるので、範囲内の見出し行すべてに配る
        for r in range(rng.min_row, rng.max_row + 1):
            if r not in header_rows:
                continue
            for c in range(rng.min_col, rng.max_col + 1):
                filled.setdefault(f"{get_column_letter(c)}{r}", val)
    return filled


def read_sheet2_map(path):
    """Sheet2 から (Sheet1セル -> (集計列番号, 旧転記名)) の対応表を作る。"""
    ws = openpyxl.load_workbook(path)["Sheet2"]
    mapping = {}
    for r in range(4, 53):
        for c in range(2, 69):
            f = ws.cell(r, c).value
            if not isinstance(f, str) or not f.startswith("=+Sheet1!"):
                continue
            src = f[len("=+Sheet1!"):]
            col_no = ws.cell(2, c).value
            legacy = norm(ws.cell(3, c).value)
            mapping[src] = (int(col_no) if col_no else None, legacy)
    return mapping


def build_headers(cache, header_rows, col):
    """見出し行を縦に連結して区分名を作る (例: 払込用紙/再発行/可 -> 払込用紙・再発行・可)。"""
    parts = []
    for r in header_rows:
        v = norm(cache.get(f"{col}{r}"))
        # C3:C5 のような縦結合は同じ値が複数行に配られるので、連続する重複は畳む
        if v and (not parts or parts[-1] != v):
            parts.append(v)
    return "・".join(parts)


def extract(form_path, print_path):
    cache = read_external_cache(form_path)
    all_header_rows = {r for b in BLOCKS for r in b[3]}
    cache = merged_fill(form_path, cache, all_header_rows)
    s2map = read_sheet2_map(form_path)

    products, kubun = [], []
    pid = kid = 0

    for bid, bname, is_product, header_rows, data_row, prod_rows, cols in BLOCKS:
        if is_product:
            order = 0
            for r in prod_rows:
                name = norm(cache.get(f"A{r}"))
                order += 1
                pid += 1
                products.append({
                    "製品ID": pid, "ブロックID": bid,
                    "製品名": name or f"（予備{order}）",
                    "表示順": order, "有効": 1 if name else 0,
                    "元セル": f"A{r}",
                })
        order = 0
        for col in cols:
            probe = f"{col}{data_row}" if data_row else f"{col}{list(prod_rows)[0]}"
            col_no, legacy = s2map.get(probe, (None, ""))
            if col_no is None:
                continue
            name = build_headers(cache, header_rows, col)
            order += 1
            kid += 1
            kubun.append({
                "区分ID": kid, "ブロックID": bid, "ブロック名": bname,
                "区分名": name or legacy or f"（未使用{order}）", "集計列ID": col_no,
                "表示順": order,
                # 入力画面に見出しが無い＝オペレータが入力できない列は無効にする
                "有効": 1 if name else 0,
                "旧転記名": legacy, "元列": col,
            })

    # ブロック7 特殊な問合せ
    for order, (name, cell) in enumerate(SPECIAL, start=1):
        kid += 1
        kubun.append({
            "区分ID": kid, "ブロックID": 7, "ブロック名": "特殊な問合せ",
            "区分名": name, "集計列ID": 7, "表示順": order,
            "有効": 1, "旧転記名": "特殊な問合せ", "元列": cell,
        })

    tasks = []
    for order, (no, label_cell, count_cell, report_name) in enumerate(TASKS, start=1):
        tasks.append({
            "業務項目ID": order, "番号": no,
            "項目名": norm(cache.get(label_cell)) or report_name,
            "帳票表示名": report_name, "件数セル": count_cell,
            "表示順": order, "有効": 1,
        })

    operators = extract_operators(form_path, print_path)
    return products, kubun, tasks, operators


def extract_operators(form_path, print_path):
    """記入用フォームのファイル名と Sheet1!B1/C1、印刷用フォームの氏名一覧から担当者を作る。"""
    folder = os.path.dirname(form_path)
    rows = []
    fullnames = {}
    wb = openpyxl.load_workbook(print_path)
    if "Sheet1" in wb.sheetnames:
        ws = wb["Sheet1"]
        for r in range(6, 21):
            v = norm(ws.cell(r, 17).value)          # Q列 = 氏名一覧
            if v:
                fullnames[v.split(" ")[0]] = v

    for fn in sorted(os.listdir(folder)):
        m = re.match(r"^(\d{3})(.+?)記入用フォーム\.xlsm$", fn)
        if not m:
            continue
        code, label = m.group(1), m.group(2)
        ws = openpyxl.load_workbook(os.path.join(folder, fn))["Sheet1"]
        sei, mei = norm(ws["B1"].value), norm(ws["C1"].value)
        staff = "職員" if code == "100" else "パート"
        # ファイル名の氏と Sheet1!B1 が食い違うファイルは、ファイル名側を正とする。
        # 顧客Ｇ/顧客G のような全半角ゆれは食い違いとみなさない。
        mismatch = 0 if fold(sei) == fold(label) or label.startswith(sei) else 1
        mei_ok = "" if mismatch else mei
        # 氏名は 印刷用フォームの氏名一覧 > フォーム自身の B1+C1 > ファイル名 の順で採用する
        own = f"{sei} {mei}".strip() if not mismatch and sei else ""
        full = fullnames.get(label) or own or label
        rows.append({
            "担当者ID": len(rows) + 1, "担当者コード": code, "姓": label,
            "名": mei_ok, "氏名": full,
            "職員区分": staff, "表示順": len(rows) + 1,
            "有効": 0 if label == "予備" else 1,
            "旧B1値": sei, "名前設定ミス": mismatch,
        })
    return rows


def write_csv(path, rows, fields):
    with open(path, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)
    print(f"  {path}  ({len(rows)} 行)")


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--src", required=True, help="現行 xlsm 一式が入ったフォルダ")
    ap.add_argument("--out", required=True, help="出力先 (data/ の親フォルダ)")
    args = ap.parse_args()

    form = os.path.join(args.src, "012髙橋記入用フォーム.xlsm")
    printf = os.path.join(args.src, "日報集計印刷用フォーム(前営業日）.xlsm")
    for p in (form, printf):
        if not os.path.exists(p):
            sys.exit(f"見つかりません: {p}")

    products, kubun, tasks, operators = extract(form, printf)
    data_dir = os.path.join(args.out, "data")
    os.makedirs(data_dir, exist_ok=True)

    write_csv(os.path.join(data_dir, "M_集計列.csv"),
              [{"集計列ID": i, "集計列名": n, "表示順": k}
               for k, (i, n) in enumerate(SHUKEI_COLUMNS, start=1)],
              ["集計列ID", "集計列名", "表示順"])
    write_csv(os.path.join(data_dir, "M_ブロック.csv"),
              [{"ブロックID": b[0], "ブロック名": b[1], "製品別": int(b[2]), "表示順": k}
               for k, b in enumerate(BLOCKS, start=1)]
              + [{"ブロックID": 7, "ブロック名": "特殊な問合せ", "製品別": 0, "表示順": 7}],
              ["ブロックID", "ブロック名", "製品別", "表示順"])
    write_csv(os.path.join(data_dir, "M_製品.csv"), products,
              ["製品ID", "ブロックID", "製品名", "表示順", "有効", "元セル"])
    write_csv(os.path.join(data_dir, "M_区分.csv"), kubun,
              ["区分ID", "ブロックID", "ブロック名", "区分名", "集計列ID",
               "表示順", "有効", "旧転記名", "元列"])
    write_csv(os.path.join(data_dir, "M_業務項目.csv"), tasks,
              ["業務項目ID", "番号", "項目名", "帳票表示名", "件数セル", "表示順", "有効"])
    write_csv(os.path.join(data_dir, "M_担当者.csv"), operators,
              ["担当者ID", "担当者コード", "姓", "名", "氏名", "職員区分",
               "表示順", "有効", "旧B1値", "名前設定ミス"])


if __name__ == "__main__":
    main()
